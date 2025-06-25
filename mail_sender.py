import yagmail
from openpyxl import load_workbook


SENDER_ID = 'rakshitcosmo13@gmail.com'
APP_PASSWORD = 'whsc ahjv iidk xwup'
EXCEL_FILE = 'emails.xlsx'
SUBJECT = 'Trial run'
BODY = 'This is an automated email sent via Python.'


def get_list(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    emails = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        email = row[0]
        if email:
            emails.append(email)
    return emails


def send_emails(email_list):
    yag = yagmail.SMTP(user=SENDER_ID, password=APP_PASSWORD)
    for email in email_list:
        try:
            yag.send(to=email, subject=SUBJECT, contents=BODY)
            print(f"Email sent to {email}")
        except Exception as e:
            print(f"Failed to send to {email}: {e}")


if __name__ == '__main__':
    email_list = get_list(EXCEL_FILE)
    print(f"Found {len(email_list)} email(s). Sending...")
    send_emails(email_list)
